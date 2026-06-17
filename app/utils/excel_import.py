"""
Excel import utilities for data import operations.
Provides reusable functions for importing student data and OCR ledger data.
"""
import io
import os
import re
import zipfile
from datetime import datetime
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.utils import get_column_letter
from app import db
from app.models import Member, Book, BookCopy
from app.models.member import generate_member_id


# Configuration
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
UPLOAD_FOLDER = 'uploads/imports'

# Default form levels
DEFAULT_FORM_LEVELS = {
    'Form 1': 1,
    'Form 2': 2,
    'Form 3': 3,
    'Form 4': 4,
    'Form 5': 5,
    'Graduated': 6,
}

# Default class groups (easily modifiable)
DEFAULT_CLASS_GROUPS = [
    'Science 1',
    'Science 2',
    'Arts 1',
    'Arts 2',
    'Commerce',
    'Technical',
]


# Header tokens that mark a sheet as having a structured header row (as opposed
# to a plain class roster where row 1 is the class name and column A is names).
HEADER_TOKENS = {
    'name', 'full_name', 'fullname', 'nama', 'student_name',
    'member_id', 'memberid', 'id', 'email', 'phone', 'no',
    'class', 'class_group', 'kelas', 'form', 'form_level', 'tingkatan',
}


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _sanitize_xlsx_bytes(raw):
    """
    Repair common openpyxl-incompatible quirks in real-world .xlsx files.

    Files exported by some Excel/LibreOffice versions (e.g. with the
    "Arial Narrow" font) set <family val="34"/> in xl/styles.xml. openpyxl
    rejects any font family > 14 with a hard ValueError, so the whole upload
    fails to load. We clamp those out-of-range values so the data is readable.

    Returns a BytesIO of the patched workbook, or None if it isn't a zip.
    """
    try:
        zin = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile:
        return None

    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == 'xl/styles.xml':
                text = data.decode('utf-8', 'replace')
                text = re.sub(
                    r'<family val="(\d+)"\s*/>',
                    lambda m: '<family val="2"/>' if int(m.group(1)) > 14 else m.group(0),
                    text,
                )
                data = text.encode('utf-8')
            zout.writestr(item, data)
    out.seek(0)
    return out


def _safe_load_workbook(filepath, **kwargs):
    """
    Load an .xlsx workbook, transparently repairing files that openpyxl's
    strict validation would otherwise reject. Falls back to the sanitized
    copy only when the normal load fails.
    """
    try:
        return openpyxl.load_workbook(filepath, **kwargs)
    except Exception:
        with open(filepath, 'rb') as fh:
            raw = fh.read()
        patched = _sanitize_xlsx_bytes(raw)
        if patched is None:
            raise
        return openpyxl.load_workbook(patched, **kwargs)


def _clean(value):
    """Normalise a cell value to a trimmed string (or '')."""
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _parse_class_title(text):
    """
    Parse a class header like "1 BESTARI", "1B", "TINGKATAN 1 SETIA" or
    "5 Science 1" into (form_level, class_name).

    Returns (form_level or None, class_name title-cased). The class name keeps
    any trailing number (e.g. "Science 1") but drops the leading form digit.
    """
    raw = _clean(text)
    if not raw:
        return None, None

    # Drop a leading "Tingkatan"/"Form" word if present
    cleaned = re.sub(r'^\s*(tingkatan|form|ting)\s*', '', raw, flags=re.IGNORECASE)

    form_level = None
    name = cleaned
    m = re.match(r'^\s*(\d)\s*(.*)$', cleaned)
    if m:
        form_level = int(m.group(1))
        remainder = m.group(2).strip()
        # "1B" (no space) -> form 1, class "B"; "1 BESTARI" -> form 1, "Bestari"
        name = remainder if remainder else cleaned[1:].strip()

    name = name.strip(' -:')
    if not name:
        return form_level, None
    # Title-case alphabetic class names but leave short codes (e.g. "B") alone
    pretty = name.title() if any(c.isalpha() for c in name) and len(name) > 2 else name
    return form_level, pretty


def save_upload_file(file):
    """Save uploaded file to uploads/imports folder"""
    if not file or file.filename == '':
        return None, "No file selected"
    
    if not allowed_file(file.filename):
        return None, "File type not allowed. Use Excel (.xlsx, .xls) or CSV"
    
    if len(file.read()) > MAX_FILE_SIZE:
        file.seek(0)
        return None, "File size exceeds 10MB limit"
    
    file.seek(0)
    
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    filename = secure_filename(file.filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
    filepath = os.path.join(UPLOAD_FOLDER, timestamp + filename)
    file.save(filepath)
    
    return filepath, None


def read_excel_file(filepath):
    """Read Excel file and return list of dictionaries"""
    try:
        workbook = _safe_load_workbook(filepath, data_only=True)
        sheet = workbook.active

        # Read header row (first row)
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
        
        if not headers:
            return None, "Excel file has no headers in first row"
        
        # Read data rows
        rows = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            row_data = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    header = headers[col_idx]
                    # Preserve original value for non-string cells
                    value = cell.value
                    if value is not None:
                        row_data[header] = str(value).strip() if isinstance(value, str) else value
                    else:
                        row_data[header] = None
            
            if any(row_data.values()):  # Skip empty rows
                row_data['_row_number'] = row_idx
                rows.append(row_data)
        
        workbook.close()
        return rows, None
    
    except Exception as e:
        return None, f"Error reading Excel file: {str(e)}"


def _row_values(row):
    """Return a list of cleaned string values for a row of cells."""
    return [_clean(c.value) for c in row]


def parse_student_workbook(filepath):
    """
    Parse a student Excel workbook into a normalised, preview-friendly structure
    that works for BOTH layouts we see in the field:

    1. Class-roster layout (SMK Munshi "SENARAI NAMA" files): one worksheet per
       class, row 1 is the class title (e.g. "1 BESTARI"), and column A lists
       the student names. There is NO header row.

    2. Structured layout: a header row (full_name/name/email/class/form ...) with
       one student per row. Works whether the data is in one sheet or many.

    Returns: (sheets, error) where sheets is a list of dicts:
        {
          'sheet': <worksheet name>,
          'format': 'roster' | 'columns',
          'form_level': <int|None>,        # suggested default for the sheet
          'class_group': <str|None>,       # suggested default for the sheet
          'students': [
             {'full_name', 'form_level', 'class_group', 'email', 'phone'}
          ],
        }
    """
    try:
        workbook = _safe_load_workbook(filepath, data_only=True)
    except Exception as e:
        return None, f"Could not open Excel file: {str(e)}"

    sheets = []
    try:
        for ws in workbook.worksheets:
            rows = [r for r in ws.iter_rows(values_only=False)]
            if not rows:
                continue

            first_row_vals = [v for v in _row_values(rows[0]) if v]
            lowered = {v.lower() for v in first_row_vals}
            is_structured = bool(lowered & HEADER_TOKENS)

            if is_structured:
                sheet_dict = _parse_columns_sheet(ws.title, rows)
            else:
                sheet_dict = _parse_roster_sheet(ws.title, rows)

            if sheet_dict and sheet_dict['students']:
                sheets.append(sheet_dict)
    finally:
        workbook.close()

    if not sheets:
        return None, "No student names found in the file."

    return sheets, None


def _parse_roster_sheet(sheet_title, rows):
    """Parse a single class-roster worksheet (class title in row 1, names below)."""
    title_text = next((v for v in _row_values(rows[0]) if v), '')
    form_level, class_name = _parse_class_title(title_text)
    # Fall back to the sheet tab name (e.g. "1B") when row 1 has no class title
    if not class_name:
        form_level, class_name = _parse_class_title(sheet_title)

    students = []
    for row in rows[1:]:
        # Take the first non-empty cell in the row as the student's name
        name = next((v for v in _row_values(row) if v), '')
        if not name:
            continue
        # Skip a repeated class-title or obvious header line
        if _parse_class_title(name)[1] and name.upper() == title_text.upper():
            continue
        students.append({
            'full_name': name,
            'form_level': form_level,
            'class_group': class_name,
            'email': None,
            'phone': None,
        })

    return {
        'sheet': sheet_title,
        'format': 'roster',
        'form_level': form_level,
        'class_group': class_name,
        'students': students,
    }


def _parse_columns_sheet(sheet_title, rows):
    """Parse a worksheet that has a structured header row."""
    headers = [v.lower() for v in _row_values(rows[0])]

    def find(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    name_idx = find('full_name', 'fullname', 'name', 'nama', 'student_name')
    email_idx = find('email')
    phone_idx = find('phone')
    class_idx = find('class_group', 'class', 'kelas')
    form_idx = find('form_level', 'form', 'tingkatan')

    students = []
    seen_form, seen_class = None, None
    for row in rows[1:]:
        vals = _row_values(row)

        def at(idx):
            return vals[idx] if idx is not None and idx < len(vals) else ''

        name = at(name_idx) if name_idx is not None else next((v for v in vals if v), '')
        if not name:
            continue

        form_level = None
        form_raw = at(form_idx)
        if form_raw:
            digits = ''.join(filter(str.isdigit, form_raw))
            if digits:
                form_level = max(1, min(6, int(digits)))
        class_name = at(class_idx) or None

        seen_form = seen_form or form_level
        seen_class = seen_class or class_name

        students.append({
            'full_name': name,
            'form_level': form_level,
            'class_group': class_name,
            'email': at(email_idx) or None,
            'phone': at(phone_idx) or None,
        })

    return {
        'sheet': sheet_title,
        'format': 'columns',
        'form_level': seen_form,
        'class_group': seen_class,
        'students': students,
    }


def commit_student_records(records):
    """
    Create Member rows from a list of normalised student dicts (typically the
    edited preview coming back from the React import wizard).

    Each record: {full_name, form_level, class_group, email, phone, member_type}

    Returns: (success_count, errors, imported_members)
    """
    success_count = 0
    errors = []
    imported = []
    default_password = 'Munshi123'

    for i, rec in enumerate(records, start=1):
        try:
            full_name = _clean(rec.get('full_name'))
            if not full_name:
                errors.append(f"Row {i}: Missing name")
                continue

            form_level = rec.get('form_level')
            try:
                form_level = max(1, min(6, int(form_level))) if form_level not in (None, '') else 1
            except (ValueError, TypeError):
                form_level = 1

            class_group = _clean(rec.get('class_group')) or None
            email = _clean(rec.get('email')) or None
            phone = _clean(rec.get('phone')) or None
            member_type = _clean(rec.get('member_type')) or 'Student'

            # Skip duplicate emails to avoid unique-constraint failures
            if email and Member.query.filter_by(email=email).first():
                errors.append(f"Row {i}: Email '{email}' already exists — skipped")
                continue

            member = Member(
                member_id=generate_member_id(),
                full_name=full_name,
                email=email,
                phone=phone,
                member_type=member_type,
                form_level=form_level,
                class_group=class_group,
                student_year=datetime.now().year,
                is_active=True,
            )
            member.set_password(default_password)
            db.session.add(member)
            db.session.flush()  # allocate id + surface errors per-row

            success_count += 1
            imported.append({
                'member_id': member.member_id,
                'full_name': full_name,
                'form_level': form_level,
                'class_group': class_group,
            })
        except Exception as e:
            db.session.rollback()
            errors.append(f"Row {i} ({rec.get('full_name', '?')}): {str(e)}")
            continue

    if success_count > 0:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            errors.insert(0, f"Database commit error: {str(e)}")
            return 0, errors, []

    # Make sure any newly seen classes are remembered for the dropdown
    try:
        _remember_classes(imported)
    except Exception:
        db.session.rollback()

    return success_count, errors, imported


def _remember_classes(imported):
    """Persist any new class names from an import into the ClassGroup table."""
    from app.models import ClassGroup
    existing = {c.name.lower() for c in ClassGroup.query.all()}
    added = False
    for rec in imported:
        name = (rec.get('class_group') or '').strip()
        if name and name.lower() not in existing:
            db.session.add(ClassGroup(name=name, form_level=rec.get('form_level'), is_active=True))
            existing.add(name.lower())
            added = True
    if added:
        db.session.commit()


def import_student_data(filepath, class_mapping=None):
    """
    Import student data from Excel file.
    
    Expected columns (case-insensitive):
    - full_name or name (required)
    - member_id (optional - auto-generated if empty)
    - email (optional)
    - phone (optional)
    - form_level or form (optional, default: 1)
    - class_group or class (optional)
    - member_type (optional, default: 'Student')
    
    Returns: (success_count, error_list, imported_members)
    """
    rows, error = read_excel_file(filepath)
    if error:
        return 0, [error], []
    
    if not rows:
        return 0, ["Excel file contains no data rows"], []
    
    success_count = 0
    errors = []
    imported_members = []
    
    for row in rows:
        try:
            row_number = row.pop('_row_number')
            
            # Validate required fields
            full_name = row.get('full_name') or row.get('name')
            if not full_name:
                errors.append(f"Row {row_number}: Missing 'full_name' or 'name'")
                continue
            
            # Get or generate member_id
            member_id = row.get('member_id', '').strip()
            if not member_id:
                member_id = generate_member_id()
            
            # Check for duplicate member_id
            existing = Member.query.filter_by(member_id=member_id).first()
            if existing:
                errors.append(f"Row {row_number}: Member ID '{member_id}' already exists")
                continue
            
            # Parse form level
            form_level_str = row.get('form_level') or row.get('form', '1')
            form_level = 1
            if isinstance(form_level_str, str):
                if 'form' in form_level_str.lower():
                    try:
                        form_level = int(''.join(filter(str.isdigit, form_level_str)))
                    except ValueError:
                        form_level = 1
                else:
                    try:
                        form_level = int(form_level_str)
                    except ValueError:
                        form_level = 1
            else:
                try:
                    form_level = int(form_level_str)
                except (ValueError, TypeError):
                    form_level = 1
            
            form_level = max(1, min(6, form_level))  # Clamp 1-6
            
            # Get class group
            class_group = row.get('class_group') or row.get('class')
            
            # Create member
            member = Member(
                member_id=member_id,
                full_name=full_name.strip(),
                email=row.get('email', '').strip() or None,
                phone=row.get('phone', '').strip() or None,
                member_type=row.get('member_type', 'Student').strip(),
                form_level=form_level,
                class_group=class_group.strip() if class_group else None,
                student_year=datetime.now().year,
                is_active=True
            )
            
            db.session.add(member)
            db.session.flush()  # Flush to get the ID
            
            success_count += 1
            imported_members.append({
                'member_id': member_id,
                'full_name': full_name,
                'form_level': form_level,
                'class_group': class_group,
                'id': member.id
            })
        
        except Exception as e:
            errors.append(f"Row {row_number}: {str(e)}")
            db.session.rollback()
            continue
    
    if success_count > 0:
        try:
            db.session.commit()
        except Exception as e:
            errors.insert(0, f"Database commit error: {str(e)}")
            db.session.rollback()
            success_count = 0
            imported_members = []
    
    return success_count, errors, imported_members


def import_ocr_ledger_data(filepath, book_id=None):
    """
    Import OCR ledger data from Excel file.
    Maps the 7-column Malaysian ledger format to book records.
    
    7 Main Columns (from Malaysian library ledgers):
    1. NO PEROLEHAN (Accession Number) - unique identifier
    2. NO PANGGILAN (Call Number) - library classification
    3. PENGARANG (Author) - book author(s)
    4. TAJUK BUKU (Title) - book title
    5. TEMPAT DAN NAMA PENERBIT (Place & Publisher) - publication location & publisher
    6. TARIKH PENERBIT (Publication Date/Year) - year of publication
    7. PUNCA (Source/Acquisition) - how book was acquired
    
    Optional fields: TARIKH PEROLEHAN (Acquisition Date), BIL NO, HARGA RM/SEN (Price),
    MUKA SURAT (Page Count), CATATAN (Notes)
    
    Expected column names (case-insensitive, supports English and Malay):
    - title, tajuk_buku, tajuk (required)
    - author, pengarang (optional)
    - publisher, penerbit, tempat_dan_nama_penerbit (optional)
    - year, tarikh_penerbit, publication_year, tahun_penerbit (optional)
    - isbn (optional)
    - accession_number, no_perolehan (optional - used for book copy)
    - call_number, no_panggilan (optional)
    - acquisition_date, tarikh_perolehan (optional)
    - source, punca (optional)
    - notes, catatan (optional)
    - price_rm, harga_rm, price (optional)
    - page_count, muka_surat (optional)
    
    Returns: (success_count, error_list, imported_books)
    """
    rows, error = read_excel_file(filepath)
    if error:
        return 0, [error], []
    
    if not rows:
        return 0, ["Excel file contains no data rows"], []
    
    success_count = 0
    errors = []
    imported_books = []
    
    for row in rows:
        try:
            row_number = row.pop('_row_number')
            
            # ===== REQUIRED: Book Title =====
            # Check multiple column name variations
            title = (row.get('title') or row.get('tajuk_buku') or 
                    row.get('tajuk') or row.get('book title'))
            if not title:
                errors.append(f"Row {row_number}: Missing book title (title/tajuk_buku/tajuk)")
                continue
            
            # Check for duplicate title
            existing_book = Book.query.filter_by(title=title.strip()).first()
            if existing_book:
                # Could update existing, but for now skip to prevent duplicates
                errors.append(f"Row {row_number}: Book title '{title[:50]}' already exists")
                continue
            
            # ===== Create new book record =====
            book = Book()
            book.title = title.strip()
            
            # ===== COLUMN 3: PENGARANG (Author) =====
            book.author = (row.get('author') or row.get('pengarang') or 
                          row.get('authorname') or '').strip() or None
            
            # ===== COLUMN 5: PENERBIT (Publisher - Place & Name) =====
            # Try multiple column name variations
            publisher = (row.get('publisher') or row.get('penerbit') or
                        row.get('tempat_dan_nama_penerbit') or row.get('publisher_name') or '').strip()
            if publisher:
                book.publisher = publisher
            
            # ===== COLUMN 6: TARIKH PENERBIT (Publication Year/Date) =====
            year_str = (row.get('year') or row.get('tarikh_penerbit') or 
                       row.get('publication_year') or row.get('tahun_penerbit') or
                       row.get('pub_year') or '')
            if year_str:
                try:
                    # Try to extract just the year if it's part of a larger date
                    year_match = re.search(r'\b(19\d{2}|20\d{2})\b', str(year_str))
                    if year_match:
                        book.publication_year = int(year_match.group(1))
                    else:
                        book.publication_year = int(year_str)
                except (ValueError, TypeError, AttributeError):
                    # Year parsing failed, skip it
                    pass
            
            # ISBN (optional)
            book.isbn = (row.get('isbn') or row.get('isbn_number') or '').strip() or None
            
            # Page count (optional)
            pages_str = row.get('page_count') or row.get('muka_surat') or row.get('pages')
            if pages_str:
                try:
                    book.page_count = int(pages_str)
                except (ValueError, TypeError):
                    pass
            
            # Category (default to 'General' if not provided)
            book.category = (row.get('category') or row.get('kategori') or
                            row.get('subject') or 'General').strip()

            # COLUMN 2: NO PANGGILAN (Call Number) - lives on Book, not BookCopy
            book.call_number = (str(row.get('call_number') or row.get('no_panggilan') or
                               row.get('callnumber') or '').strip()[:32]) or None
            
            # Notes/Remarks (optional)
            description = (row.get('notes') or row.get('catatan') or 
                          row.get('remarks') or row.get('description') or '').strip()
            if description:
                book.description = description
            
            # Price (optional)
            price_str = row.get('price_rm') or row.get('harga_rm') or row.get('price')
            if price_str:
                try:
                    # Extract numeric price value
                    price_match = re.search(r'(\d+)[.,]?(\d{0,2})', str(price_str))
                    if price_match:
                        book.price = float(f"{price_match.group(1)}.{price_match.group(2) or '0'}")
                except (ValueError, TypeError, AttributeError):
                    pass
            
            db.session.add(book)
            db.session.flush()  # Get the book ID
            
            # ===== Create BookCopy with accession information =====
            # COLUMN 1: NO PEROLEHAN (Accession Number)
            accession_num = (row.get('accession_number') or row.get('no_perolehan') or 
                            row.get('accession_num') or '').strip()
            
            if accession_num:
                # COLUMN 7: PUNCA (Source - how acquired)
                source = (row.get('source') or row.get('punca') or 
                         row.get('acquisition_source') or '').strip() or 'Purchase'
                
                # Tarikh Perolehan (Acquisition Date)
                acq_date_str = (row.get('acquisition_date') or row.get('tarikh_perolehan') or
                               row.get('acq_date') or '')
                acq_date = None
                if acq_date_str:
                    try:
                        # Try to parse various date formats
                        from datetime import datetime as dt
                        acq_date = dt.strptime(str(acq_date_str), '%Y-%m-%d').date()
                    except:
                        try:
                            acq_date = dt.strptime(str(acq_date_str), '%d/%m/%Y').date()
                        except:
                            pass  # Skip invalid dates
                
                copy = BookCopy(
                    book_id=book.id,
                    accession_number=accession_num[:32],
                    barcode=accession_num.replace('-', '').replace(' ', '').upper()[:32],
                    status='available',
                    acquisition_source=source,
                    acquisition_date=acq_date,
                    condition='Good',  # Default condition
                    location='Main Stack'  # Default location
                )
                db.session.add(copy)
            
            success_count += 1
            imported_books.append({
                'title': title,
                'author': book.author,
                'publisher': book.publisher,
                'year': book.publication_year,
                'isbn': book.isbn,
                'accession': accession_num if accession_num else 'N/A',
                'book_id': book.id
            })
        
        except Exception as e:
            errors.append(f"Row {row_number}: {str(e)}")
            db.session.rollback()
            continue
    
    if success_count > 0:
        try:
            db.session.commit()
        except Exception as e:
            errors.insert(0, f"Database commit error: {str(e)}")
            db.session.rollback()
            success_count = 0
            imported_books = []
    
    return success_count, errors, imported_books


def get_class_groups():
    """
    Get the list of class names for the dropdown.

    Merges three sources, de-duplicated and sorted:
      1. Admin-managed ClassGroup table (seeded with SMK Munshi streams)
      2. Distinct class_group values already used by members
      3. The hard-coded fallback defaults (in case the table is empty)
    """
    names = {}  # lower -> display, preserves first-seen casing
    try:
        from app.models import ClassGroup
        for cg in ClassGroup.query.filter_by(is_active=True).all():
            if cg.name:
                names.setdefault(cg.name.lower(), cg.name)
        rows = db.session.query(Member.class_group).distinct().filter(
            Member.class_group.isnot(None), Member.class_group != ''
        ).all()
        for (name,) in rows:
            if name and name.strip():
                names.setdefault(name.strip().lower(), name.strip())
    except Exception:
        pass

    if not names:
        return [c.upper() for c in DEFAULT_CLASS_GROUPS]

    # Classes are displayed UPPERCASE per school policy; de-dup case-insensitively.
    return sorted({v.upper() for v in names.values()})


def get_form_levels():
    """Get list of form levels"""
    return [
        {'value': 1, 'label': 'Form 1'},
        {'value': 2, 'label': 'Form 2'},
        {'value': 3, 'label': 'Form 3'},
        {'value': 4, 'label': 'Form 4'},
        {'value': 5, 'label': 'Form 5'},
        {'value': 6, 'label': 'Graduated'},
    ]
